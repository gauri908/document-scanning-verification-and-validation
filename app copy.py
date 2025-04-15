from flask import Flask, request, render_template, redirect, url_for, session, send_from_directory, Response, json, jsonify
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
import os
import time
from datetime import datetime
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import io
# Local modules (assumed to exist or provided)
from modules.ocr_extractor import extract_text_from_file
from modules.database_manager import save_student_data
from modules.field_extractor import extract_fields
from modules.summary_generator import generate_summary_report

# --- App Config ---
app = Flask(__name__)
app.secret_key = os.urandom(24)  # Secure random key
UPLOAD_FOLDER = 'database'
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max upload size
USER_DB = 'users.json'

DOCUMENT_FIELDS = {
    "adhar_card": "Aadhar Card",
    "pan_card": "PAN Card",
    "ssc_certificate": "SSC Certificate",
    "hsc_certificate": "HSC Certificate",
    "cast_certificate": "Caste Certificate",
    "cast_validity": "Caste Validity",
    "non_cremelier": "Non-Creamy Layer Certificate",
    "domicile": "Domicile Certificate",
    "school_leaving": "School Leaving Certificate"
}

# --- Utility Functions ---
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def load_users():
    if os.path.exists(USER_DB):
        with open(USER_DB, 'r') as f:
            return json.load(f)
    return {}

def save_users(users):
    with open(USER_DB, 'w') as f:
        json.dump(users, f, indent=2)

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'username' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper

def get_available_students():
    upload_folder = app.config['UPLOAD_FOLDER']
    if not os.path.exists(upload_folder):
        return []
    students = []
    for folder in os.listdir(upload_folder):
        if os.path.isdir(os.path.join(upload_folder, folder)):
            try:
                parts = folder.split('_')
                if len(parts) == 4:
                    name, father_name, surname, dob = parts
                    full_name = f"{name} {father_name} {surname}"
                    students.append({
                        'folder': folder,
                        'display': f"{full_name} — {dob[:2]}/{dob[2:4]}/{dob[4:]}"
                    })
            except ValueError:
                continue
    return students

# --- Routes ---
@app.route('/index', methods=['GET', 'POST'])
@login_required
def index():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        father_name = request.form.get('father_name', '').strip()
        grandfather_name = request.form.get('grandfather_name', '').strip()
        surname = request.form.get('surname', '').strip()
        dob = request.form.get('dob', '').strip()  # YYYY-MM-DD from date input
        gender = request.form.get('gender', '').strip()
        pincode = request.form.get('pincode', '').strip()

        # Convert DOB to DD/MM/YYYY
        try:
            dob_dt = datetime.strptime(dob, "%Y-%m-%d")
            dob_formatted = dob_dt.strftime("%d/%m/%Y")
        except ValueError:
            dob_formatted = "NA"

        folder_name = f"{name}_{father_name}_{surname}_{dob_formatted.replace('/', '')}"
        session['student_folder'] = folder_name
        session['student_fullname'] = f"{name} {father_name} {surname}"

        save_student_data(name, father_name, grandfather_name, surname, dob_formatted, gender, pincode)
        return redirect(url_for('upload_documents'))
    return render_template('index.html')

@app.route('/upload', methods=['GET', 'POST'])
@login_required
def upload_documents():
    folder_name = session.get('student_folder')
    full_name = session.get('student_fullname')
    if folder_name and not os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], folder_name)):
        session.pop('student_folder', None)
        session.pop('student_fullname', None)
        folder_name = None

    if request.method == 'POST' and folder_name:
        folder_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)
        os.makedirs(folder_path, exist_ok=True)
        for field_key in DOCUMENT_FIELDS:
            file = request.files.get(field_key)
            if file and allowed_file(file.filename):
                ext = file.filename.rsplit('.', 1)[1].lower()
                filename = f"{field_key}.{ext}"
                file_path = os.path.join(folder_path, secure_filename(filename))
                file.save(file_path)
        return redirect(url_for('validate_documents'))

    uploaded_files = {}
    if folder_name:
        folder_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)
        for key in DOCUMENT_FIELDS:
            for ext in ALLOWED_EXTENSIONS:
                check_path = os.path.join(folder_path, f"{key}.{ext}")
                if os.path.exists(check_path):
                    uploaded_files[key] = f"{key}.{ext}"
                    break

    return render_template('upload.html',
                           full_name=full_name,
                           folder=folder_name,
                           doc_fields=DOCUMENT_FIELDS,
                           uploaded_files=uploaded_files)

@app.route('/validate-page')
@login_required
def validate_documents():  # Renamed for consistency
    folder_name = session.get('student_folder')
    if folder_name and not os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], folder_name)):
        session.pop('student_folder', None)
        session.pop('student_fullname', None)
        return redirect(url_for('index'))
    return render_template('validate.html', doc_fields=DOCUMENT_FIELDS)

@app.route('/view/<folder>/<filename>')
@login_required
def view_file(folder, filename):
    return send_from_directory(os.path.join(app.config['UPLOAD_FOLDER'], folder), filename)

@app.route('/validate')
@login_required
def validate():
    folder_name = session.get('student_folder')
    if not folder_name:
        return "No folder selected", 400

    folder_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)

    def generate():
        for filename in os.listdir(folder_path):
            if filename.lower().endswith(('.pdf', '.png', '.jpg', '.jpeg')):
                base_name = os.path.splitext(filename)[0]
                txt_path = os.path.join(folder_path, f"{base_name}.txt")
                if os.path.exists(txt_path):
                    yield f"data:{json.dumps({'doc': base_name, 'status': 'skipped', 'message': 'Already exists'})}\n\n"
                    continue
                file_path = os.path.join(folder_path, filename)
                try:
                    message = extract_text_from_file(file_path, folder_path)
                    yield f"data:{json.dumps({'doc': base_name, 'status': 'success', 'message': message})}\n\n"
                except Exception as e:
                    yield f"data:{json.dumps({'doc': base_name, 'status': 'error', 'message': str(e)})}\n\n"
                time.sleep(0.5)
        yield f"data:{json.dumps({'status': 'done'})}\n\n"
    return Response(generate(), mimetype='text/event-stream')

@app.route('/extract-fields')
@login_required
def extract_fields_route():
    folder_name = session.get('student_folder')
    if not folder_name:
        return "No folder selected", 400

    folder_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)
    all_fields_path = os.path.join(folder_path, "all_fields.json")

    if os.path.exists(all_fields_path):
        with open(all_fields_path, "r", encoding="utf-8") as f:
            all_data = json.load(f)
    else:
        all_data = {}

    def generate():
        for filename in os.listdir(folder_path):
            if filename.lower().endswith('.txt'):
                doc_name = os.path.splitext(filename)[0]
                if doc_name in all_data:
                    yield f"data:{json.dumps({'doc': doc_name, 'status': 'skipped', 'message': 'Already extracted'})}\n\n"
                    continue
                file_path = os.path.join(folder_path, filename)
                try:
                    result = extract_fields(file_path)
                    if "error" in result:
                        yield f"data:{json.dumps({'doc': doc_name, 'status': 'error', 'message': result['error']})}\n\n"
                    else:
                        all_data[doc_name] = result
                        with open(all_fields_path, "w", encoding="utf-8") as f:
                            json.dump(all_data, f, indent=4, ensure_ascii=False)
                        yield f"data:{json.dumps({'doc': doc_name, 'status': 'success', 'message': 'Fields extracted'})}\n\n"
                except Exception as e:
                    yield f"data:{json.dumps({'doc': doc_name, 'status': 'error', 'message': str(e)})}\n\n"
                time.sleep(0.5)
        yield f"data:{json.dumps({'status': 'done'})}\n\n"
    return Response(generate(), mimetype='text/event-stream')

@app.route('/check-validation-status')
@login_required
def check_validation_status():
    folder_name = session.get('student_folder')
    if not folder_name or not os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], folder_name)):
        return jsonify({'status': 'needs_upload'})

    folder_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)
    all_fields_path = os.path.join(folder_path, "all_fields.json")
    has_documents = any(filename.lower().endswith(('.pdf', '.png', '.jpg', '.jpeg')) 
                       for filename in os.listdir(folder_path))
    
    if not has_documents:
        return jsonify({'status': 'needs_upload'})
    elif not os.path.exists(all_fields_path):
        return jsonify({'status': 'needs_validation'})
    else:
        return jsonify({'status': 'validated'})

@app.route('/result')
@login_required
def result_page():
    folder_name = session.get('student_folder')
    if not folder_name or not os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], folder_name)):
        session.pop('student_folder', None)
        session.pop('student_fullname', None)
        return render_template("result.html", student=None, extracted_data={}, user_input={}, pending_step=None)

    folder_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)
    all_fields_path = os.path.join(folder_path, "all_fields.json")
    user_input_path = os.path.join(folder_path, "student_data.json")

    extracted_data = {}
    user_input = {}
    pending_step = None

    if not os.path.exists(user_input_path):
        pending_step = "upload"
    elif not os.path.exists(all_fields_path):
        pending_step = "validate"
    else:
        with open(all_fields_path, "r", encoding="utf-8") as f:
            extracted_data = json.load(f)
        if os.path.exists(user_input_path):
            with open(user_input_path, "r", encoding="utf-8") as f:
                user_input = json.load(f)

    return render_template("result.html", 
                          student=folder_name, 
                          extracted_data=extracted_data, 
                          user_input=user_input, 
                          pending_step=pending_step,
                          doc_fields=DOCUMENT_FIELDS)

@app.route('/summary')
@login_required
def summary_report():
    folder_name = session.get('student_folder')
    if not folder_name or not os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], folder_name)):
        session.pop('student_folder', None)
        session.pop('student_fullname', None)
        return redirect(url_for('index'))

    folder_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)
    user_input_path = os.path.join(folder_path, "student_data.json")
    all_fields_path = os.path.join(folder_path, "all_fields.json")

    if not os.path.exists(user_input_path) or not os.path.exists(all_fields_path):
        return "Summary cannot be generated. Please complete previous steps."

    summary = generate_summary_report(user_input_path, all_fields_path)
    return render_template("summary.html", summary_text=summary, student=folder_name)

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        confirm_password = request.form['confirm_password']

        if password != confirm_password:
            return render_template('signup.html', error="Passwords do not match.")
        users = load_users()
        if username in users:
            return render_template('signup.html', error="User already exists.")
        users[username] = generate_password_hash(password)
        save_users(users)
        return redirect(url_for('login'))
    return render_template('signup.html')

@app.route('/')
def landing():
    return render_template('landing.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        users = load_users()
        if username in users and check_password_hash(users[username], password):
            session['username'] = username
            return redirect(url_for('index'))
        return render_template('login.html', error="Invalid username or password.")
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('username', None)
    session.pop('student_folder', None)
    session.pop('student_fullname', None)
    return redirect(url_for('login'))

@app.route('/get-students', methods=['GET'])
@login_required
def get_students():
    students = get_available_students()
    return jsonify(students)

@app.route('/select-student', methods=['POST'])
@login_required
def select_student():
    data = request.get_json()
    student_folder = data.get('student_folder')
    if student_folder and os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], student_folder)):
        session['student_folder'] = student_folder
        parts = student_folder.split('_')
        if len(parts) == 4:
            name, father_name, surname, _ = parts
            session['student_fullname'] = f"{name} {father_name} {surname}"
        return jsonify({'success': True})
    return jsonify({'success': False}), 400

# @app.route('/check-extraction-status')
# @login_required
# def check_extraction_status():
#     folder_name = session.get('student_folder')
#     if not folder_name:
#         return jsonify({'text': {}, 'fields': {}})
#     folder_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)
#     text_status = {}
#     field_status = {}
#     for doc_key in DOCUMENT_FIELDS.keys():  # Only check relevant documents
#         text_file = os.path.join(folder_path, f"{doc_key}.txt")
#         field_file = os.path.join(folder_path, f"{doc_key}.json")
#         text_status[doc_key] = os.path.exists(text_file)
#         field_status[doc_key] = os.path.exists(field_file)
#     return jsonify({'text': text_status, 'fields': field_status})

@app.route('/check-extraction-status')
@login_required
def check_extraction_status():
    folder_name = session.get('student_folder')
    if not folder_name:
        return jsonify({'text': {}, 'fields': {}})
    
    folder_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)
    all_fields_path = os.path.join(folder_path, "all_fields.json")
    
    text_status = {}
    field_status = {}
    
    # Load all_fields.json if it exists
    all_fields_data = {}
    if os.path.exists(all_fields_path):
        with open(all_fields_path, 'r', encoding='utf-8') as f:
            all_fields_data = json.load(f)

    # Check status for each document in DOCUMENT_FIELDS
    for doc_key in DOCUMENT_FIELDS.keys():
        # Check for text extraction (.txt file exists)
        text_file = os.path.join(folder_path, f"{doc_key}.txt")
        text_status[doc_key] = os.path.exists(text_file) and os.path.getsize(text_file) > 0  # Ensure file isn’t empty
        
        # Check for field extraction (entry exists in all_fields.json)
        field_status[doc_key] = doc_key in all_fields_data and bool(all_fields_data[doc_key])  # Ensure data isn’t empty
    
    return jsonify({'text': text_status, 'fields': field_status})


@app.route('/export')
@login_required
def export_to_excel():
    folder_name = session.get('student_folder')
    if not folder_name or not os.path.exists(os.path.join(app.config['UPLOAD_FOLDER'], folder_name)):
        session.pop('student_folder', None)
        session.pop('student_fullname', None)
        return redirect(url_for('result_page'))

    folder_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)
    all_fields_path = os.path.join(folder_path, "all_fields.json")
    user_input_path = os.path.join(folder_path, "student_data.json")

    if not os.path.exists(all_fields_path) or not os.path.exists(user_input_path):
        return "Cannot export: Missing data.", 400

    with open(all_fields_path, "r", encoding="utf-8") as f:
        extracted_data = json.load(f)
    with open(user_input_path, "r", encoding="utf-8") as f:
        user_input = json.load(f)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Verification Results"

    # Define headers
    fields = ["Field", "User Input"] + [DOCUMENT_FIELDS[doc] for doc in DOCUMENT_FIELDS.keys()]
    ws.append(fields)

    # Define data rows
    rows = [
        ["Name", user_input.get('name', '-')],
        ["Father's Name", user_input.get('father_name', '-')],
        ["Grandfather's Name", user_input.get('grandfather_name', '-')],
        ["Surname", user_input.get('surname', '-')],
        ["Date of Birth", user_input.get('dob', '-')],
        ["Gender", user_input.get('gender', '-')],
        ["Pincode", user_input.get('pincode', '-')]
    ]

    # Define field mappings for extracted data
    field_mappings = {
        "Name": "name",
        "Father's Name": "father name",
        "Grandfather's Name": "grandfather name",
        "Surname": "surname",
        "Date of Birth": "DOB",
        "Gender": "gender",
        "Pincode": "pincode"
    }

    # Fill colors
    green_fill = PatternFill(start_color="2F855A", end_color="2F855A", fill_type="solid")  # Match
    red_fill = PatternFill(start_color="C53030", end_color="C53030", fill_type="solid")    # No match

    # Populate rows with extracted data and apply coloring
    for row_idx, (field, user_value) in enumerate(rows, start=2):
        row_data = [field, user_value]
        field_key = field_mappings[field]
        for doc in DOCUMENT_FIELDS.keys():
            extracted_value = extracted_data.get(doc, {}).get(field_key, '-')
            row_data.append(extracted_value)
        ws.append(row_data)

        # Apply coloring to the row
        for col_idx, doc in enumerate(DOCUMENT_FIELDS.keys(), start=3):  # Start at column C
            cell = ws.cell(row=row_idx, column=col_idx)
            extracted_value = extracted_data.get(doc, {}).get(field_key, '').lower()
            user_value_lower = user_value.lower()
            if extracted_value == user_value_lower and extracted_value != 'na' and user_value_lower != 'na':
                cell.fill = green_fill
            elif extracted_value != user_value_lower and extracted_value != 'na' and user_value_lower != 'na':
                cell.fill = red_fill

    # Save to a BytesIO object
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Send the file with the folder name
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"{folder_name}.xlsx"
    )

if __name__ == '__main__':
    app.run(debug=True)